import openpyxl
import os, sys

try:
    from global_param import Global
except ImportError:
    sys.path.append(os.path.abspath('../'))
    from ishangzu.global_param import Global

class Exporter:
    def __init__(self):
        self.wb = openpyxl.Workbook()
        self.sheets = {
            'all': self.wb.create_sheet('ALL', 0),
            'hz': self.wb.create_sheet('HZ', 1),
            'sh': self.wb.create_sheet('SH', 2),
            'nj': self.wb.create_sheet('NJ', 3),
            'sz': self.wb.create_sheet('SZ', 4)}

        self.wb.remove_sheet(self.wb.get_sheet_by_name('Sheet'))


    def insert_sheet(self, sheet_name, data):
        """
        数据插入知道sheet
        :param sheet_name: sheet名称
        :param data: 数据源list（dict)
        :return:
        """
        sheet = self.sheets[sheet_name]
        sheet.append(['daystr', 'cityname', 'mark_type', 'event_name', 'click_count'])
        for l in data:
            sheet.append([l['daystr'],l['cityname'],l['mark_type'],l['event_name'],l['click_count']])



    def get_data(self,daystr=None):
        """
        根据日期获取各个城市数据
        :param daystr:日期
        :return: 数据
        """
        sql_dict = {'all' : "select * from page_click_stat where 1=1",
                    'hz': "select * from page_click_stat where cityname='hz'",
                    'sh': "select * from page_click_stat where cityname='sh'",
                    'nj': "select * from page_click_stat where cityname='nj'",
                    'sz': "select * from page_click_stat where cityname='sz'"}

        data_dict = dict()
        for k,v in sql_dict.items():
            if daystr is None:
                data_dict[k] = Global.mysql_helper.fetchall(v+' order by daystr, cityname desc')
            else:
                data_dict[k] = Global.mysql_helper.fetchall(v+" and daystr='%s' order by daystr, cityname desc" % daystr)

        return data_dict


    def export(self,daystr=None):
        """
        导出到excel
        :param daystr:
        :return:
        """
        data_dict = self.get_data(daystr)
        for sheet_name,data in data_dict.items():
            self.insert_sheet(sheet_name,data)

        if daystr is None:
            filename = 'page_click_stat_all.xlsx'
        else:
            filename = 'page_click_stat_%s.xlsx' % daystr

        self.wb.save(filename)

    def test(self):
        sql = "select * from local_buried lb left join page_click_stat stat  where stat.cityname=''"


        sql = "select * from page_click_stat where cityname='sh' or cityname=''"
        data = Global.mysql_helper.fetchall(sql)
        self.insert_sheet('nj',data)






if __name__ == '__main__':
    exporter = Exporter()
    exporter.export(daystr='2017-04-16')