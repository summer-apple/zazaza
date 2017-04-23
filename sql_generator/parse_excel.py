from openpyxl import load_workbook
from sql_generator.mysqlconn import MySQLHelper

class ExcelParser:
    def __init__(self,db_name,filename):
        if not filename.endswith('.xls'):
            if not filename.endswith('.xlsx'):
                raise Exception('ONLY EXCEL FILE ACCEPT!')
        self.wb = load_workbook(filename=filename)
        self.db_name = db_name

        self.mysql_helper = MySQLHelper()


        # for ws in [self.wb[sn] for sn in self.wb.sheetnames]:
        #     for row in ws.rows:
        #         tmp = ''
        #         for cell in row:
        #
        #             tmp = tmp + str(cell.value)+'\t' if cell.value is not None else tmp+'None\t'
        #         print(tmp)



    def create_db(self):

        #db_name, db_desp, charset = [cell.value for cell in list(self.head.rows)[1]][:3]

        #print(db_name,db_desp,charset)
        self.mysql_helper.execute('drop database if exists %s' % self.db_name)

        sql = "create database %s charset utf8" % self.db_name
        self.mysql_helper.execute(sql)
        self.mysql_helper.close()
        self.mysql_helper = MySQLHelper(self.db_name)

    def create_table(self):
        tb_names = self.wb.sheetnames[1:]

        for tb_name, ws in [(sn,self.wb[sn]) for sn in tb_names]:

            sql = "create table %s (" % tb_name.lower()


            pks = list()

            for row in list(ws.rows)[2:]:
                column_name, data_type, length, nn, ai, pk, default, comment, uindex1, uindex2, uindex3, index1, index2, index3 = [cell.value for cell in row]
                print(column_name, data_type, length, nn, ai, pk, default, comment, uindex1, uindex2, uindex3, index1, index2, index3)




                if column_name is not None and column_name != 'BACK_TO_HEAD':
                    sql = sql + column_name.lower() +' '

                    # length
                    if data_type.lower() in ('char','varchar','decimal'):
                        sql = sql +data_type+ '('+str(length)+') '
                    else:
                        sql = sql + data_type + ' '

                    # default
                    if nn is not None:
                        sql = sql + 'not null '

                    # auto increase
                    if ai is not None:
                        sql = sql + 'auto_increment '

                    # default
                    if default is not None:
                        if data_type.lower() in ['int','float','double']:
                            sql = sql + "default " + str(default)+" "
                        else:
                            sql = sql + "default '" + str(default)+"' "


                    # comment
                    if comment is not None:
                        sql = sql + "comment '"+ comment+"'"

                    sql = sql + ','

                    # pk
                    if pk is not None:
                        pks.append(column_name)
                        #sql = sql + 'primary key '

            if len(pks)>0:
                sql = sql+'primary key (%s)' % str(pks)[1:-1].replace("'","")+','

            sql = sql[:-1]+")"

            print(sql)
            self.mysql_helper.execute(sql)

if __name__ == '__main__':
    ep = ExcelParser('db_chaoyi','sql.xlsx')
    ep.create_db()
    ep.create_table()