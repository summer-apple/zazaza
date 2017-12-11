import openpyxl
from openpyxl import load_workbook
import numpy as np
import pandas as pd
from matplotlib import pyplot as plt
from pyspark import SparkContext, SparkConf, SQLContext
from pyspark.sql import DataFrameReader, DataFrameWriter
from pyspark.sql.types import StructType,StringType,IntegerType,StructField
import fire

class Analysiser:
    def __init__(self):

        conf = SparkConf().setAppName('Analysiser').set("spark.sql.crossJoin.enabled", True)
        self.sc = SparkContext(conf=conf)
        self.sqlctx = SQLContext(self.sc)

        self.pdf = pd.read_excel('data_o.xlsx', sheetname=0, header=0,  parse_cols=[9, 10, 23, 32, 45, 60])

        schema = StructType([
            StructField('TI',StringType(),True),
            StructField('SO', StringType(), True),
            StructField('C1', StringType(), True),
            StructField('TC', StringType(), True),
            StructField('PY', StringType(), True),
            StructField('UT', StringType(), True)

        ])

        df = self.sqlctx.createDataFrame(self.pdf,schema)

        def m_clean(x):
            try:
                py = int(x['PY'])
                tc = int(x['TC'])
                authors = x['C1']

                if py>=2006 and py<=2016 and authors != '':

                    first_author = authors[1:].split(']')[0].split('; ')[0]

                    return [(x['TI'],x['SO'],x['C1'],first_author,x['TC'],int(x['PY']),x['UT']),]
                else:
                    return []
            except Exception as e:
                return []

        schema2 = StructType([
            StructField('TI', StringType(), True),
            StructField('SO', StringType(), True),
            StructField('C1', StringType(), True),
            StructField('first_author', StringType(), True),
            StructField('TC', StringType(), True),
            StructField('PY', IntegerType(), True),
            StructField('UT', StringType(), True)

        ])
        self.df = self.sqlctx.createDataFrame(df.rdd.flatMap(m_clean),schema2)


        #self.df.show()


    # def parse(self):
    #     .wb = load_workbook('data_min.xlsx')
    #     sheet = wb.get_sheet_by_name('all')
    #     new_wb = openpyxl.Workbook()
    #     new_sheet = new_wb.create_sheet('simple')
    #     new_sheet.append(['TI', 'SO', 'C1', 'TC', 'PY', 'UT'])
    #
    #
    #     for row in list(sheet.rows)[2:100]:
    #         r = [c.value for c in row]
    #         r_min = [r[9],r[10],r[23],r[32],r[45],r[60]]
    #         print(r_min)
    #         new_sheet.append(r_min)
    #     new_wb.save('export.xlsx')

    def parse2(self):

        self.df.ExcelWriter('output.xls')





    def func1(self):
        df = self.df.toPandas()
        #print(df.head())
        plt.figure(figsize=(9, 6))
        plt.scatter(df['PY'], df['TC'], s=25, alpha=0.4, marker='o')
        # T:散点的颜色
        # s：散点的大小
        # alpha:是透明程度
        plt.show()


    def func2(self):
        df = self.df
        first_author_df = df.select('first_author','PY').groupBy('first_author').max('PY').withColumnRenamed('max(PY)','maxPY')

        self.sqlctx.registerDataFrameAsTable(df.drop('first_author'),'df')
        self.sqlctx.registerDataFrameAsTable(first_author_df,'fa')

        sql = "select first_author,TC from (fa outer join df on C1 like CONCAT('%',first_author,'%'))"

        join = self.sqlctx.sql(sql)
        join_rdd = join.rdd.map(lambda x:(x['first_author'],x['TC'])).reduceByKey(lambda x,y:x+'-'+y)

        # for r in join_rdd.collect():
        #     print(r)

        def m_h(x):
            flag = False
            h = 0
            cts = [int(x) for x in x[1].split('-')]
            cts.sort(reverse=True)
            for i in range(1, len(list(cts))+1):
                if i >= cts[i-1]:
                    flag = True
                    h = i # TODO or cts[i-1]
                    break

            if flag:
                return [(x[0],h),]
            else:
                return []

        author_h_rdd = join_rdd.flatMap(m_h)
        author_h_df = self.sqlctx.createDataFrame(author_h_rdd,['first_author','h'])
        final_df = author_h_df.join(first_author_df,'first_author','left_outer').select('h','maxPY')
        pdf = final_df.toPandas()

        plt.figure(figsize=(9, 6))
        plt.scatter(pdf['maxPY'], pdf['h'], s=25, alpha=0.4, marker='o')
        # T:散点的颜色
        # s：散点的大小
        # alpha:是透明程度
        plt.show()


if __name__ == '__main__':
    fire.Fire(Analysiser)

    #a.fun1()